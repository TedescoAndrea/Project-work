import requests
import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Lista degli URL da cui fare scraping
urls = [
    "https://cir-reports.cir-safety.org/FetchCIRReports",
    "https://cir-reports.cir-safety.org/FetchCIRReports/?&pagingcookie=%26lt%3bcookie%20page%3d%26quot%3b1%26quot%3b%26gt%3b%26lt%3bpcpc_name%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_ingredientidname%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_cirrelatedingredientsid%20last%3d%26quot%3b%7bC223037E-F278-416D-A287-2007B9671D0C%7d%26quot%3b%20first%3d%26quot%3b%7b940AF697-52B5-4A3A-90A6-B9DB30EF4A7E%7d%26quot%3b%20%2f%26gt%3b%26lt%3b%2fcookie%26gt%3b&page=2",
]

def fetch_cir_reports(url):
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        print("Errore durante il recupero dei dati:", response.status_code)
        return None

def extract_data_from_json(data):
    records = []
    if data:
        results = data.get("results", [])
        for result in results:
            ingredient_name = result.get("pcpc_ingredientname", "")
            inci_name = result.get("pcpc_ciringredientname", "")
            id_link = result.get("pcpc_ingredientid", "")
            link = f"https://cir-reports.cir-safety.org/cir-ingredient-status-report/?id={id_link}"
            records.append({
                "Ingredient_Name": ingredient_name,
                "INCI_Name": inci_name,
                "Link": link
            })
    else:
        print("Nessun dato disponibile")
    return records

def fetch_and_extract():
    all_records = []
    for url in urls:
        print(f"Recupero dati dall'URL: {url}")
        cir_data = fetch_cir_reports(url)
        if cir_data:
            records = extract_data_from_json(cir_data)
            all_records.extend(records)
    return all_records

def update_excel_file(new_records, file_path):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        ws.delete_rows(2, ws.max_row)
        wb.save(file_path)
        new_df = pd.DataFrame(new_records)
        new_df.drop_duplicates(inplace=True)
        new_df.to_excel(file_path, index=False)
        print(f"Dati aggiornati in '{file_path}'")
    else:
        new_df = pd.DataFrame(new_records)
        new_df.drop_duplicates(inplace=True)
        new_df.to_excel(file_path, index=False)
        print(f"Dati salvati in '{file_path}'")

def remove_first_row_if_needed(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # Controlla se la prima cella della prima riga contiene "Ingredient_Name"
    if ws.cell(row=1, column=1).value == "Ingredient_Name":
        ws.delete_rows(1, 1)
        wb.save(file_path)
        print(f"Prima riga rimossa da '{file_path}'")
    else:
        print("La prima riga non contiene 'Ingredient_Name', nessuna riga rimossa.")

def remove_duplicates_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Convertiamo i dati del foglio in un DataFrame pandas
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)

    # Rimuoviamo i duplicati basandoci sulla colonna "Ingredient_Name"
    df.drop_duplicates(subset=["Ingredient_Name"], keep="first", inplace=True)

    # Riscriviamo i dati nel foglio Excel
    ws.delete_rows(2, ws.max_row)  # Rimuoviamo tutte le righe tranne l'intestazione
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(file_path)
    print(f"Dati duplicati rimossi da '{file_path}'")


def miao():
    file_path = "cir_reports.xlsx"
    print("Recupero i dati da CIR...")
    new_records = fetch_and_extract()
    if new_records:
        update_excel_file(new_records, file_path)
        print("Aggiornamento completato")
        remove_duplicates_excel(file_path)
        remove_first_row_if_needed(file_path)
    else:
        print("Nessun nuovo dato da aggiornare")



